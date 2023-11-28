import copy

from openpyxl import utils
from openpyxl.cell import Cell
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


class CopyXlsx:
    """复制xlsx"""

    @classmethod
    def copy_sheet(cls, worksheet: Worksheet, copy_worksheet: Worksheet):
        """复制excel的sheet并带上相应格式"""
        # 1) 复制excel的宽与高
        cls._copy_widths(worksheet, copy_worksheet)
        cls._copy_heights(worksheet, copy_worksheet)
        # 2) 复制excel中的数值
        all_row_cells = worksheet.iter_rows()
        for row, row_cells in enumerate(all_row_cells):
            for col, cell in enumerate(row_cells):
                copy_cell = worksheet.cell(row + 1, col + 1, cell.value)
                if isinstance(cell, MergedCell):
                    continue  # excel中有合并了的单元格，先跳过，复制其他单元格的格式
                cls._copy_cell_format(cell, copy_cell)
        # 3) 复制excel合并单元格
        cls._copy_merged_cell(worksheet, copy_worksheet)

    @staticmethod
    def _copy_heights(worksheet: Worksheet, copy_worksheet: Worksheet):
        """复制excel的高"""
        for row in range(worksheet.max_row):
            copy_worksheet.row_dimensions[row].height = worksheet.row_dimensions[row].height

    @staticmethod
    def _copy_widths(worksheet: Worksheet, copy_worksheet: Worksheet):
        """复制excel的宽"""
        for col in range(worksheet.max_column):
            col_letter = utils.get_column_letter(col + 1)
            copy_worksheet.column_dimensions[col_letter].width = worksheet.column_dimensions[col_letter].width

    @staticmethod
    def _copy_cell_format(cell: Cell, copy_cell: Cell):
        """复制excel的单元格的格式"""
        copy_cell.data_type = cell.data_type
        copy_cell.fill = copy.copy(cell.fill)
        if cell.has_style:
            # copy_cell._style = copy.copy(cell._style)
            copy_cell.font = copy.copy(cell.font)
            copy_cell.border = copy.copy(cell.border)
            copy_cell.fill = copy.copy(cell.fill)
            copy_cell.number_format = copy.copy(cell.number_format)
            copy_cell.protection = copy.copy(cell.protection)
            copy_cell.alignment = copy.copy(cell.alignment)
        if cell.hyperlink:
            copy_cell._hyperlink = copy.copy(cell.hyperlink)
        if cell.comment:
            copy_cell.comment = copy.copy(cell.comment)

    @staticmethod
    def _copy_merged_cell(worksheet: Worksheet, copy_worksheet: Worksheet):
        """复制excel合并单元格"""
        # 开始处理合并单元格形式为“(<CellRange A1:A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        for index, cell in enumerate(list(worksheet.merged_cells)):
            merge_cell = str(cell).replace('(<CellRange ', '').replace('>,)', '')
            copy_worksheet.merge_cells(merge_cell)
