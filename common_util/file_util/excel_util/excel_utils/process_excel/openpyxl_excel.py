import re
import typing

from openpyxl import styles, utils
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class OpenpyxlExcel:
    _height_difference = 6  # 这个值表示字体大小与行高差，约定为6

    @classmethod
    def add_title(cls, worksheet: Worksheet, title: str, length: int, title_size: float):
        """添加标题"""
        cls.insert_row_data(worksheet, [title], 1)
        # 设置标题单元格合并居中加粗
        cls.merge_cells(worksheet, "A1", f"{utils.get_column_letter(length)}1")
        cls.set_cell(worksheet['A1'], horizontal='center', vertical='center', is_border=True,
                     font_size=title_size, is_bold=True)
        worksheet.row_dimensions[1].height = title_size + cls._height_difference

    @classmethod
    def add_tag(cls, worksheet: Worksheet, tags: typing.List[str], tag_row: int, tag_size: float, tag_color: str):
        """添加表头"""
        cls.insert_row_data(worksheet, tags, tag_row)
        for cell in worksheet[tag_row]:
            cls.set_cell(cell, horizontal='center', vertical='center', is_border=True,
                         font_size=tag_size, is_bold=True)
            if cell.value is not None:  # 当表头为空时，不设置这个单元格背景颜色
                cls.set_cell(cell, fill_color=tag_color)
        worksheet.row_dimensions[tag_row].height = tag_size + cls._height_difference

    @classmethod
    def format_row_style(cls, worksheet: Worksheet, row: int, data_size: int):
        """格式化行样式"""
        worksheet.row_dimensions[row].height = data_size + cls._height_difference  # 设置行高
        for cell in worksheet[row]:
            vertical = "center" if cell.alignment.vertical is None else cell.alignment.vertical
            cls.set_cell(cell, vertical=vertical, is_border=True)

    @staticmethod
    def get_sheet_heights(worksheet: Worksheet) -> typing.List[float]:
        """获取sheet行高"""
        return [worksheet.row_dimensions[row + 1].height for row in range(worksheet.max_row)]

    @staticmethod
    def get_sheet_widths(worksheet: Worksheet) -> typing.List[float]:
        """获取sheet列宽"""
        return [worksheet.column_dimensions[utils.get_column_letter(col + 1)].width for col in
                range(worksheet.max_column)]

    @classmethod
    def insert_row_data(cls, worksheet: Worksheet, data_list: list, row: int):
        """插入一条数据"""
        # 1) 不指定row时，将数据写入文件最底部
        if row is None:
            worksheet.append(data_list)
        # 2) 在指定行插入一条数据
        else:
            worksheet.insert_rows(row)  # 在第row行上方插入一行空行
            for index, data in enumerate(data_list):
                worksheet.cell(row, index + 1, data)
            # 使用insert方法插入一行后，excel会多出一行可以被获取的空行，因此在插入之后需要检测一下并删除，不然使用append方法添加数据会有空行
            if not any([cell.value for cell in worksheet[row + 1]]):
                worksheet.delete_rows(row + 1)

    @staticmethod
    def merge_cells(worksheet: Worksheet, coordinate_from: str, coordinate_to: str):
        """合并单元格"""
        # 处理起始坐标与结束坐标，使起始坐标必须在结束坐标左上
        from_col, from_row = [each for each in re.compile(r'(\d+|\s+)').split(coordinate_from) if each]
        to_col, to_row = [each for each in re.compile(r'(\d+|\s+)').split(coordinate_to) if each]
        new_from_col, new_to_col = (from_col, to_col) if from_col < to_col else (to_col, from_col)
        new_from_row, new_to_row = (from_row, to_row) if int(from_row) < int(to_row) else (to_row, from_row)
        # 合并单元格
        worksheet.merge_cells(f"{new_from_col}{new_from_row}:{new_to_col}{new_to_row}")

    @classmethod
    def set_cell(cls, cell: Cell, value: any = '_no_value', number_format: str = None,
                 horizontal: typing.Optional[str] = '_no_horizontal', vertical: typing.Optional[str] = '_no_vertical',
                 fill_color: typing.Optional[str] = '_no_fill_color', is_border: bool = None,
                 font_size: float = None, is_bold: bool = None, is_italic: bool = None,
                 font_color: typing.Optional[str] = '_no_font_color'
                 ):
        """设置单元格"""
        # 1) 设置单元格的值
        if value != "_no_value":
            cell.value = value
        # 2) 设置单元格格式
        if number_format is not None:
            cell.number_format = number_format
        # 2) 设置单元格对齐
        cls._set_cell_alignment(cell, horizontal, vertical)
        # 3) 设置单元格填充色
        cls._set_cell_fill(cell, fill_color)
        # 4) 设置单元格边框
        if is_border is not None:
            side = styles.Side(style='thin', color='000000') if is_border else None  # 设置边框属性
            cell.border = styles.Border(left=side, right=side, top=side, bottom=side)
        # 5) 设置单元格字体
        cls._set_cell_font(cell, font_size, is_bold, is_italic, font_color)

    @staticmethod
    def set_sheet_heights(worksheet: Worksheet, heights: typing.List[typing.Optional[float]]):
        """设置sheet行高"""
        for index, height in enumerate(heights):
            if height is None:  # 当设置行高为None时，不修改列值
                continue
            worksheet.row_dimensions[index + 1].height = height

    @staticmethod
    def set_sheet_widths(worksheet: Worksheet, widths: typing.List[typing.Optional[float]]):
        """设置sheet列宽"""
        for index, width in enumerate(widths):
            if width is None:  # 当设置列宽为None时，不修改列值
                continue
            worksheet.column_dimensions[utils.get_column_letter(index + 1)].width = width

    @staticmethod
    def _set_cell_alignment(cell: Cell, horizontal: str, vertical: str):
        """设置单元格对齐"""
        if horizontal not in ["left", "right", "center", None]:
            horizontal = cell.alignment.horizontal
        if vertical not in ["bottom", "top", "center", None]:
            vertical = cell.alignment.vertical
        if horizontal == cell.alignment.horizontal and vertical == cell.alignment.vertical:
            return
        cell.alignment = styles.Alignment(horizontal=horizontal, vertical=vertical)

    @staticmethod
    def _set_cell_fill(cell: Cell, fill_color: typing.Optional[str]):
        """设置单元格颜色"""
        if fill_color == "_no_fill_color":
            pass
        elif fill_color is None:
            cell.fill = styles.fills.PatternFill(fill_type=None)
        else:
            cell.fill = styles.fills.PatternFill(fill_type='solid', fgColor=fill_color)

    @staticmethod
    def _set_cell_font(cell: Cell, font_size: float, is_bold: bool, is_italic: bool, font_color: typing.Optional[str]):
        """设置单元格字体"""
        if [font_size, is_bold, is_italic, font_color] == [None, None, None, "_no_font_color"]:
            return
        # 1.1) 设置字体大小
        font_size = font_size if font_size is not None else cell.font.size
        # 1.2) 设置字体是否加粗
        is_bold = is_bold if is_bold is not None else cell.font.bold
        # 1.3) 设置字体是否倾斜
        is_italic = is_italic if is_italic is not None else cell.font.italic
        # 1.4) 设置字体颜色
        font_color = font_color if font_color != "_no_font_color" else cell.font.color
        # 2) 设置字体
        cell.font = styles.Font(size=font_size, bold=is_bold, italic=is_italic, color=font_color)
