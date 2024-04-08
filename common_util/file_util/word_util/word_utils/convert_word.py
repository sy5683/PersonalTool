import logging
import os
import typing
from pathlib import Path

import docx
import openpyxl
from win32com.client import constants, gencache


class ConvertWord:
    """转换word"""

    @staticmethod
    def word_to_excel(word_path: str, save_path: typing.Union[Path, str]) -> str:
        """word转excel"""
        logging.info(f"开始将Word文件转换为Excel: {word_path}")
        save_path = f"{os.path.splitext(word_path)[0]}.xlsx" if save_path is None else str(save_path)
        assert not os.path.exists(save_path), f"文件已存在，无法转换: {save_path}"
        document = docx.Document(word_path)
        workbook = openpyxl.Workbook()
        for index, table in enumerate(document.tables):
            worksheet = workbook.active if not index else workbook.create_sheet()
            for row_index, row in enumerate(table.rows):
                for col_index, cell in enumerate(row.cells):
                    worksheet.cell(row_index + 1, col_index + 1, cell.text)
        workbook.save(save_path)
        workbook.close()
        logging.info(f"成功将Word文件转换为Excel: {save_path}")
        return save_path

    @staticmethod
    def word_to_pdf(word_path: str, save_path: typing.Union[Path, str]) -> str:
        """word转pdf"""
        logging.info(f"开始将word文件转换为pdf: {word_path}")
        save_path = f"{os.path.splitext(word_path)[0]}.pdf" if save_path is None else str(save_path)
        assert not os.path.exists(save_path), f"文件已存在，无法转换: {save_path}"
        app = gencache.EnsureDispatch('Word.Application')
        document = app.Documents.Open(word_path, ReadOnly=True)
        document.ExportAsFixedFormat(save_path, constants.wdExportFormatPDF, Item=constants.wdExportDocumentWithMarkup,
                                     CreateBookmarks=constants.wdExportCreateHeadingBookmarks)
        document.Close()
        app.Quit(constants.wdDoNotSaveChanges)
        logging.info(f"成功将word文件转换为pdf: {word_path}")
        return save_path
