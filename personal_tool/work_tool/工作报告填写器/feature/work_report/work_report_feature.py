import logging
import os
import typing
from pathlib import Path

import openpyxl

from common_util.file_util.file_util.file_util import FileUtil
from .entity.weekly_report import WeeklyReport
from .factory.daily_report_factory import DailyReportFactory


class WorkReportFeature:
    __work_report_path = None

    @classmethod
    def get_weekly_report(cls, target_date: str) -> WeeklyReport:
        """获取周报"""
        for weekly_report in reversed(cls.get_weekly_reports()):
            if target_date in weekly_report.date_range:
                weekly_report.set_completion_rates()  # 生成日报的完成率
                return weekly_report
        raise AttributeError(f"未找到目标日期所在的周报: {target_date}")

    @classmethod
    def get_weekly_reports(cls) -> typing.List[WeeklyReport]:
        daily_report_path = cls._get_work_report_path()
        wb = openpyxl.load_workbook(daily_report_path)
        weekly_reports = []
        weekly_report = None
        for rows in list(wb.active.iter_rows())[1:]:  # 去除表头
            row_values = [row.value for row in rows]
            if not any(row_values):  # 去除间隔行
                continue
            if row_values[1] == "周计划":
                weekly_report = WeeklyReport(row_values[2])
                weekly_reports.append(weekly_report)
            else:
                if weekly_report is None:
                    continue
                daily_report = DailyReportFactory.row_values_to_daily_report(row_values)
                weekly_report.add_daily_report(daily_report)
        wb.close()
        return weekly_reports

    @classmethod
    def _get_work_report_path(cls) -> str:
        default_work_report_path = str(Path(__file__).parent.parent.parent.joinpath("file\\日报.xlsx"))  # 工作报告默认路径
        if cls.__work_report_path is None:
            if not os.path.exists(default_work_report_path):
                logging.warning(f"默认日报文件路径不存在: {default_work_report_path}")
                cls.__work_report_path = FileUtil.get_file_path()  # 使用tkinter窗口重新获取工作报告路径
            if not cls.__work_report_path:
                cls.__work_report_path = default_work_report_path
        if not os.path.exists(cls.__work_report_path):
            FileUtil.make_dir(default_work_report_path)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(["", "日期", "今日任务", "明日计划"])
            workbook.save(cls.__work_report_path)
            workbook.close()
        return cls.__work_report_path
