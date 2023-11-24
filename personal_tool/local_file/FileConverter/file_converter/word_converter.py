import re
from pathlib import Path

import docx
import openpyxl

from .feature.file_feature import FileFeature


class WordConverter:

    @staticmethod
    def word_to_excel():
        """word转excel"""
        file_paths = FileFeature.get_file_paths()
        assert len(file_paths) == 1, "word转excel时一次只能选择一份文件"
        assert re.match(r"\.doc|\.docx", FileFeature.get_suffix()), "选择的文件无法进行word转excel操作"
        document_path = file_paths[0]
        document = docx.Document(document_path)
        wb = openpyxl.Workbook()
        for index, table in enumerate(document.tables):
            ws = wb.active if not index else wb.create_sheet()
            for row_index, row in enumerate(table.rows):
                for col_index, cell in enumerate(row.cells):
                    ws.cell(row_index + 1, col_index + 1, cell.text)
        excel_path = FileFeature.get_save_path(f"{Path(document_path).stem}.xlsx")
        wb.save(excel_path)
        wb.close()
